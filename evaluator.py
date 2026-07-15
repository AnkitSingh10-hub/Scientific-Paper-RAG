import os
from datetime import datetime
from collections import defaultdict

import gradio as gr
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

from evaluation.eval import evaluate_all_retrieval, evaluate_all_answers

load_dotenv(override=True)

# Color coding thresholds - Retrieval
MRR_GREEN = 0.9
MRR_AMBER = 0.75
NDCG_GREEN = 0.9
NDCG_AMBER = 0.75
COVERAGE_GREEN = 90.0
COVERAGE_AMBER = 75.0

# Color coding thresholds - Answer (1-5 scale)
ANSWER_GREEN = 4.5
ANSWER_AMBER = 4.0

RESULTS_DIR = "eval_results"
os.makedirs(RESULTS_DIR, exist_ok=True)


def get_color(value: float, metric_type: str) -> str:
    """Get color based on metric value and type."""
    if metric_type == "mrr":
        if value >= MRR_GREEN:
            return "green"
        elif value >= MRR_AMBER:
            return "orange"
        else:
            return "red"
    elif metric_type == "ndcg":
        if value >= NDCG_GREEN:
            return "green"
        elif value >= NDCG_AMBER:
            return "orange"
        else:
            return "red"
    elif metric_type == "coverage":
        if value >= COVERAGE_GREEN:
            return "green"
        elif value >= COVERAGE_AMBER:
            return "orange"
        else:
            return "red"
    elif metric_type in ["accuracy", "completeness", "relevance"]:
        if value >= ANSWER_GREEN:
            return "green"
        elif value >= ANSWER_AMBER:
            return "orange"
        else:
            return "red"
    return "black"


def format_metric_html(
    label: str,
    value: float,
    metric_type: str,
    is_percentage: bool = False,
    score_format: bool = False,
) -> str:
    """Format a metric with color coding."""
    color = get_color(value, metric_type)
    if is_percentage:
        value_str = f"{value:.1f}%"
    elif score_format:
        value_str = f"{value:.2f}/5"
    else:
        value_str = f"{value:.4f}"
    return f"""
    <div style="margin: 10px 0; padding: 15px; background-color: #f5f5f5; border-radius: 8px; border-left: 5px solid {color};">
        <div style="font-size: 14px; color: #666; margin-bottom: 5px;">{label}</div>
        <div style="font-size: 28px; font-weight: bold; color: {color};">{value_str}</div>
    </div>
    """


def run_retrieval_evaluation(progress=gr.Progress()):
    """Run retrieval evaluation and yield updates."""
    total_mrr = 0.0
    total_ndcg = 0.0
    total_coverage = 0.0
    category_mrr = defaultdict(list)
    details = []
    count = 0

    for test, result, prog_value in evaluate_all_retrieval():
        count += 1
        total_mrr += result.mrr
        total_ndcg += result.ndcg
        total_coverage += result.keyword_coverage

        category_mrr[test.category].append(result.mrr)

        details.append(
            {
                "Question": test.question,
                "Category": test.category,
                "MRR": result.mrr,
                "nDCG": result.ndcg,
                "Keywords Found": result.keywords_found,
                "Total Keywords": result.total_keywords,
                "Keyword Coverage (%)": result.keyword_coverage,
            }
        )

        # Update progress bar only
        progress(prog_value, desc=f"Evaluating test {count}...")

    # Calculate final averages
    avg_mrr = total_mrr / count
    avg_ndcg = total_ndcg / count
    avg_coverage = total_coverage / count

    # Create final summary metrics HTML
    final_html = f"""
    <div style="padding: 0;">
        {format_metric_html("Mean Reciprocal Rank (MRR)", avg_mrr, "mrr")}
        {format_metric_html("Normalized DCG (nDCG)", avg_ndcg, "ndcg")}
        {format_metric_html("Keyword Coverage", avg_coverage, "coverage", is_percentage=True)}
        <div style="margin-top: 20px; padding: 10px; background-color: #d4edda; border-radius: 5px; text-align: center; border: 1px solid #c3e6cb;">
            <span style="font-size: 14px; color: #155724; font-weight: bold;">✓ Evaluation Complete: {count} tests</span>
        </div>
    </div>
    """

    # Create final bar chart data
    category_data = []
    for category, mrr_scores in category_mrr.items():
        avg_cat_mrr = sum(mrr_scores) / len(mrr_scores)
        category_data.append({"Category": category, "Average MRR": avg_cat_mrr})

    df = pd.DataFrame(category_data)

    summary = {
        "Tests Run": count,
        "Average MRR": avg_mrr,
        "Average nDCG": avg_ndcg,
        "Average Keyword Coverage (%)": avg_coverage,
    }

    return final_html, df, summary, df, pd.DataFrame(details)


def run_answer_evaluation(progress=gr.Progress()):
    """Run answer evaluation and yield updates (async)."""
    total_accuracy = 0.0
    total_completeness = 0.0
    total_relevance = 0.0
    category_accuracy = defaultdict(list)
    details = []
    count = 0

    for test, result, prog_value in evaluate_all_answers():
        count += 1
        total_accuracy += result.accuracy
        total_completeness += result.completeness
        total_relevance += result.relevance

        category_accuracy[test.category].append(result.accuracy)

        details.append(
            {
                "Question": test.question,
                "Category": test.category,
                "Accuracy": result.accuracy,
                "Completeness": result.completeness,
                "Relevance": result.relevance,
                "Feedback": result.feedback,
            }
        )

        # Update progress bar only
        progress(prog_value, desc=f"Evaluating test {count}...")

    # Calculate final averages
    avg_accuracy = total_accuracy / count
    avg_completeness = total_completeness / count
    avg_relevance = total_relevance / count

    # Create final summary metrics HTML
    final_html = f"""
    <div style="padding: 0;">
        {format_metric_html("Accuracy", avg_accuracy, "accuracy", score_format=True)}
        {format_metric_html("Completeness", avg_completeness, "completeness", score_format=True)}
        {format_metric_html("Relevance", avg_relevance, "relevance", score_format=True)}
        <div style="margin-top: 20px; padding: 10px; background-color: #d4edda; border-radius: 5px; text-align: center; border: 1px solid #c3e6cb;">
            <span style="font-size: 14px; color: #155724; font-weight: bold;">✓ Evaluation Complete: {count} tests</span>
        </div>
    </div>
    """

    # Create final bar chart data
    category_data = []
    for category, accuracy_scores in category_accuracy.items():
        avg_cat_accuracy = sum(accuracy_scores) / len(accuracy_scores)
        category_data.append(
            {"Category": category, "Average Accuracy": avg_cat_accuracy}
        )

    df = pd.DataFrame(category_data)

    summary = {
        "Tests Run": count,
        "Average Accuracy": avg_accuracy,
        "Average Completeness": avg_completeness,
        "Average Relevance": avg_relevance,
    }

    return final_html, df, summary, df, pd.DataFrame(details)


# --- Excel export -----------------------------------------------------------


def _style_header_row(worksheet, n_cols: int, row: int = 1):
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, n_cols + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit_columns(worksheet, df: pd.DataFrame, start_row: int = 1):
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()]
        )
        worksheet.column_dimensions[
            worksheet.cell(row=start_row, column=i).column_letter
        ].width = min(max(max_len + 2, 10), 60)


def save_results_to_excel(
    embedding_model,
    chunking_method,
    chunk_size,
    chunk_overlap,
    vector_db,
    generation_model,
    judge_model,
    run_notes,
    retrieval_summary,
    retrieval_category_df,
    retrieval_details_df,
    answer_summary,
    answer_category_df,
    answer_details_df,
):
    """Bundle current config + whichever results exist into one xlsx file."""

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filepath = os.path.join(RESULTS_DIR, f"eval_results_{timestamp}.xlsx")

    config_rows = [
        {"Setting": "Run Timestamp", "Value": timestamp},
        {"Setting": "Embedding Model", "Value": embedding_model},
        {"Setting": "Chunking Method", "Value": chunking_method},
        {"Setting": "Chunk Size", "Value": chunk_size},
        {"Setting": "Chunk Overlap", "Value": chunk_overlap},
        {"Setting": "Vector DB", "Value": vector_db},
        {"Setting": "Generation Model", "Value": generation_model},
        {"Setting": "Judge Model", "Value": judge_model},
        {"Setting": "Notes", "Value": run_notes},
    ]

    if retrieval_summary:
        config_rows.append({"Setting": "", "Value": ""})
        config_rows.append({"Setting": "--- Retrieval Results ---", "Value": ""})
        for k, v in retrieval_summary.items():
            config_rows.append({"Setting": k, "Value": v})

    if answer_summary:
        config_rows.append({"Setting": "", "Value": ""})
        config_rows.append({"Setting": "--- Answer Results ---", "Value": ""})
        for k, v in answer_summary.items():
            config_rows.append({"Setting": k, "Value": v})

    config_df = pd.DataFrame(config_rows)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        config_df.to_excel(writer, sheet_name="Config & Summary", index=False)

        if retrieval_details_df is not None and not retrieval_details_df.empty:
            retrieval_details_df.to_excel(
                writer, sheet_name="Retrieval Details", index=False
            )
        if retrieval_category_df is not None and not retrieval_category_df.empty:
            retrieval_category_df.to_excel(
                writer, sheet_name="Retrieval by Category", index=False
            )

        if answer_details_df is not None and not answer_details_df.empty:
            answer_details_df.to_excel(writer, sheet_name="Answer Details", index=False)
        if answer_category_df is not None and not answer_category_df.empty:
            answer_category_df.to_excel(
                writer, sheet_name="Answer by Category", index=False
            )

        # Style headers + autofit on every sheet written
        for sheet_name, df in [
            ("Config & Summary", config_df),
            ("Retrieval Details", retrieval_details_df),
            ("Retrieval by Category", retrieval_category_df),
            ("Answer Details", answer_details_df),
            ("Answer by Category", answer_category_df),
        ]:
            if sheet_name in writer.sheets and df is not None and not df.empty:
                ws = writer.sheets[sheet_name]
                _style_header_row(ws, len(df.columns))
                _autofit_columns(ws, df)

    return gr.update(value=filepath, visible=True), f"Saved to `{filepath}`"


def main():
    """Launch the Gradio evaluation app."""
    theme = gr.themes.Soft(font=["Inter", "system-ui", "sans-serif"])

    with gr.Blocks(title="RAG Evaluation Dashboard", theme=theme) as app:
        gr.Markdown("# 📊 RAG Evaluation Dashboard")
        gr.Markdown(
            "Evaluate retrieval and answer quality for the Insurellm RAG system"
        )

        # State to carry results between evaluation runs and the export step
        retrieval_summary_state = gr.State(None)
        retrieval_category_state = gr.State(None)
        retrieval_details_state = gr.State(None)
        answer_summary_state = gr.State(None)
        answer_category_state = gr.State(None)
        answer_details_state = gr.State(None)

        # RETRIEVAL SECTION
        gr.Markdown("## 🔍 Retrieval Evaluation")

        retrieval_button = gr.Button("Run Evaluation", variant="primary", size="lg")

        with gr.Row():
            with gr.Column(scale=1):
                retrieval_metrics = gr.HTML(
                    "<div style='padding: 20px; text-align: center; color: #999;'>Click 'Run Evaluation' to start</div>"
                )

            with gr.Column(scale=1):
                retrieval_chart = gr.BarPlot(
                    x="Category",
                    y="Average MRR",
                    title="Average MRR by Category",
                    y_lim=[0, 1],
                    height=400,
                )

        # ANSWERING SECTION
        gr.Markdown("## 💬 Answer Evaluation")

        answer_button = gr.Button("Run Evaluation", variant="primary", size="lg")

        with gr.Row():
            with gr.Column(scale=1):
                answer_metrics = gr.HTML(
                    "<div style='padding: 20px; text-align: center; color: #999;'>Click 'Run Evaluation' to start</div>"
                )

            with gr.Column(scale=1):
                answer_chart = gr.BarPlot(
                    x="Category",
                    y="Average Accuracy",
                    title="Average Accuracy by Category",
                    y_lim=[1, 5],
                    height=400,
                )

        # SAVE SECTION (config form + save button live down here)
        gr.Markdown("## 💾 Save Results")
        gr.Markdown(
            "Fill in the run config below, then save it together with whichever "
            "evaluation(s) you've run above into a single Excel file (one sheet per section)."
        )

        with gr.Accordion("⚙️ Run Configuration", open=True):
            with gr.Row():
                embedding_model_input = gr.Textbox(
                    label="Embedding Model",
                    value="text-embedding-3-small",
                )
                chunking_method_input = gr.Dropdown(
                    label="Chunking Method",
                    choices=[
                        "Fixed-size",
                        "Recursive character",
                        "Semantic",
                        "Sentence-based",
                        "Markdown-aware",
                    ],
                    value="Recursive character",
                )
            with gr.Row():
                chunk_size_input = gr.Number(label="Chunk Size (tokens)", value=500)
                chunk_overlap_input = gr.Number(
                    label="Chunk Overlap (tokens)", value=50
                )
                vector_db_input = gr.Textbox(label="Vector DB", value="vector_database")
            with gr.Row():
                generation_model_input = gr.Textbox(
                    label="Generation Model", value="DeepSeek-V3.2"
                )
                judge_model_input = gr.Textbox(
                    label="Judge Model", value="DeepSeek-V3.2"
                )
            run_notes_input = gr.Textbox(
                label="Notes",
                placeholder="e.g. baseline run, testing new chunk size...",
                lines=2,
            )

        save_button = gr.Button("Save Results to Excel", variant="secondary", size="lg")
        save_status = gr.Markdown("")
        save_file = gr.File(label="Download", visible=False)

        # Wire up the evaluations
        retrieval_button.click(
            fn=run_retrieval_evaluation,
            outputs=[
                retrieval_metrics,
                retrieval_chart,
                retrieval_summary_state,
                retrieval_category_state,
                retrieval_details_state,
            ],
        )

        answer_button.click(
            fn=run_answer_evaluation,
            outputs=[
                answer_metrics,
                answer_chart,
                answer_summary_state,
                answer_category_state,
                answer_details_state,
            ],
        )

        save_button.click(
            fn=save_results_to_excel,
            inputs=[
                embedding_model_input,
                chunking_method_input,
                chunk_size_input,
                chunk_overlap_input,
                vector_db_input,
                generation_model_input,
                judge_model_input,
                run_notes_input,
                retrieval_summary_state,
                retrieval_category_state,
                retrieval_details_state,
                answer_summary_state,
                answer_category_state,
                answer_details_state,
            ],
            outputs=[save_file, save_status],
        )

    app.launch(inbrowser=True)


if __name__ == "__main__":
    main()
